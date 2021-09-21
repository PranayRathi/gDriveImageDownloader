#Imports
import xlrd
import openpyxl

#Function imports
from helperFunction import *
from downloadImage import download_file_from_google_drive

from uploadImage import *
from variable import *
from google_drive_downloader import GoogleDriveDownloader as gdd
import time

#open workbook
wb = openpyxl.load_workbook(filename = filename)

ws_read = wb[provided_image_sheet]
ws_write =  wb[modified_image_sheet] if sheet_exists(wb) else wb.create_sheet(modified_image_sheet)

read_sheet_row_size = ws_read.max_row
read_sheet_col_size = ws_read.max_column

# Creating new sheet and writing header ["Image1, Image2", ...]

for i in range(1, read_sheet_col_size+1):
    ws_write.cell(1, i).value = ws_read.cell(1, i).value
wb.save(filename)

# Remove files from data directory
check_directory()

product_successful_count = 0
image_modified_count = 0
unsuccessful_image_count = 0

# Entire sheet
# for row in range(2, read_sheet_row_size +1):
    
# For the required range of rows 
for row in range(309, 310):
    for col in range(2, read_sheet_col_size+1):
    # for col in range(2, 4):

        print("\n #### START OF ###### ROW NO : " + str(row) + " COLOUM NO : " + str(col))
        image_link = str(ws_read.cell(row, col).value)
        file_id = checkForGoogleLink(image_link)

        if(file_id):
            lookup_result = lookupForImageId(file_id, row, col)
            if(not lookup_result and type(lookup_result) == type(False)):
                print("Same image link found over the row Discarding it")
                
            elif(not lookup_result):
                print(file_id)
                destination = './data/img_'+ str(row) + str(col) + ".png"
                download_file_from_google_drive(file_id, destination)

                file_size_check = check_file_size(destination)
                if(file_size_check):
                    dropbox_link_data = upload_image(destination)
                else:
                    dropbox_link_data = "File size is greater than 2 MB LOC : " + str(row)+ " " + str(col)
                ws_write.cell(row, col).value = dropbox_link_data
                
                image_modified_count +=1
                print(dropbox_link_data)

            else:
                found_row, found_col = lookup_result[0], lookup_result[1]
                ws_write.cell(row, col).value = str(ws_write.cell(found_row, found_col).value)

        else:
            ws_write.cell(row, col).value = image_link
            unsuccessful_image_count +=1
        wb.save(filename)
        print("#### END OF ###### ROW NO : " + str(row) + " COLOUM NO : " + str(col))
    
    product_successful_count += 1
    print(row , "Completed \n")

print("===============Report========================")
print("Product_successful_count = ", product_successful_count)
print("image_modified_count = ", image_modified_count)
print("unsuccessful_image_count = ", unsuccessful_image_count)
